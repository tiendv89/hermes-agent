"""read_uploaded_file — download and read a chat-uploaded file's content.

Reads a file that a user attached to a chat message (PDF, DOCX, XLSX, or any
text-based format).  Downloads the raw bytes from storage-service, detects the
file type from the Content-Type response header (falling back to extension-based
detection from the filename), extracts text content, and returns it so the agent
can reason about it.

Unlike images (which require a vision-capable model), file content is extracted
as text — every model can read it.
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict

from plugins.clients.storage_service_client import StorageServiceError, download_file

logger = logging.getLogger(__name__)

# Content-Type values that trigger format-specific parsers.
_PDF_CONTENT_TYPES = frozenset({"application/pdf"})
_DOCX_CONTENT_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
)
_XLSX_CONTENT_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
)

# Maximum text length before truncation (100 KB).
_MAX_TEXT_BYTES = 100 * 1024
_TRUNCATION_MARKER = "\n[... truncated ...]"


SCHEMA: Dict[str, Any] = {
    "description": (
        "Read a chat-uploaded file's text content. Pass the file_id from the "
        "message's file_ids list. Supports PDF, DOCX, XLSX, and any text-based "
        "format (txt, csv, json, yaml, source code, etc.). Returns the extracted "
        "text content, truncated at ~100 KB for large files."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "file_id": {
                "type": "string",
                "description": "The file ID from the message's file_ids list.",
            },
            "workspace_id": {
                "type": "string",
                "description": (
                    "Workspace identifier. Omit to use the current workspace from context."
                ),
            },
        },
        "required": ["file_id"],
        "additionalProperties": False,
    },
}


def _extract_text_from_pdf(data: bytes, filename: str) -> str:
    """Extract text from PDF bytes using PyPDF2."""
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text()
        if text:
            parts.append(f"--- Page {i} ---\n{text}")
    if not parts:
        return "[PDF contained no extractable text.]"
    return "\n\n".join(parts)


def _extract_text_from_docx(data: bytes, filename: str) -> str:
    """Extract text from DOCX bytes using python-docx."""
    from docx import Document as DocxDocument

    doc = DocxDocument(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return "[DOCX contained no extractable text.]"
    return "\n\n".join(paragraphs)


def _extract_text_from_xlsx(data: bytes, filename: str) -> str:
    """Extract text from XLSX bytes using openpyxl, rendering each sheet as tabular text."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = [str(cell) if cell is not None else "" for cell in row]
            # Skip entirely empty rows
            if any(cells):
                parts.append("\t".join(cells))
    wb.close()
    if len(parts) <= 1:  # Only the sheet header, no data rows
        return "[XLSX contained no extractable data.]"
    return "\n".join(parts)


def _extract_text_as_utf8(data: bytes, filename: str) -> str:
    """Decode bytes as UTF-8 text (fallback for unrecognised formats)."""
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError(f"File is not valid UTF-8 text and is not a recognised format (PDF/DOCX/XLSX): {exc}")


def _detect_parser(content_type: str, filename: str):
    """Return the appropriate extraction function based on content type.

    Falls back to extension-based detection from the filename when the
    content type is generic (application/octet-stream or missing).
    """
    ct = (content_type or "").strip().lower()

    # Primary: content-type match.
    if ct in _PDF_CONTENT_TYPES:
        return _extract_text_from_pdf
    if ct in _DOCX_CONTENT_TYPES:
        return _extract_text_from_docx
    if ct in _XLSX_CONTENT_TYPES:
        return _extract_text_from_xlsx

    # Fallback: extension-based detection when content type is generic.
    fn = (filename or "").lower()
    if fn.endswith(".pdf"):
        return _extract_text_from_pdf
    if fn.endswith(".docx"):
        return _extract_text_from_docx
    if fn.endswith(".xlsx"):
        return _extract_text_from_xlsx

    # Default: UTF-8 text.
    return _extract_text_as_utf8


def handle(file_id: str = "", workspace_id: str = "", **_: Any) -> Dict[str, Any]:
    """Download a chat-uploaded file and extract its text content.

    Returns:
        A dict with:
          - ``ok`` (bool): True on success, False on error.
          - ``filename`` (str): The original filename (when available).
          - ``content_type`` (str): The detected content type.
          - ``text`` (str): The extracted text content.
          - ``truncated`` (bool): True if the text was truncated.
    """
    from ..context import get_org_id, get_user_id, get_workspace_id

    wid = workspace_id or get_workspace_id()
    caller_user_id = get_user_id()
    caller_org_id = get_org_id()

    if not wid:
        return {
            "ok": False,
            "error": "workspace_id is required but was not provided and no context is set.",
        }
    if not file_id or not file_id.strip():
        return {"ok": False, "error": "file_id is required."}

    try:
        result = download_file(
            wid, file_id, user_id=caller_user_id, org_id=caller_org_id
        )
    except StorageServiceError as exc:
        logger.warning("read_uploaded_file: storage-service error for %s: %s", file_id, exc)
        return {"ok": False, "error": str(exc)}
    except Exception as exc:
        logger.warning("read_uploaded_file: unexpected download error for %s: %s", file_id, exc)
        return {"ok": False, "error": str(exc)}

    data = result["data"]
    content_type = result["content_type"]
    filename = result.get("filename", "") or file_id

    # Pick the right extractor.
    try:
        extractor = _detect_parser(content_type, filename)
    except Exception:
        logger.warning("read_uploaded_file: could not determine parser for %s", file_id)
        return {"ok": False, "error": "Could not determine how to read this file type."}

    # Extract text.
    try:
        text = extractor(data, filename)
    except ValueError as exc:
        logger.warning("read_uploaded_file: extraction failed for %s: %s", file_id, exc)
        return {"ok": False, "error": str(exc)}
    except Exception as exc:
        logger.warning("read_uploaded_file: unexpected extraction error for %s: %s", file_id, exc)
        return {"ok": False, "error": f"Failed to read file: {exc}"}

    # Truncation safeguard.
    truncated = False
    if len(text.encode("utf-8")) > _MAX_TEXT_BYTES:
        # Truncate at the byte boundary.
        encoded = text.encode("utf-8")[:_MAX_TEXT_BYTES]
        text = encoded.decode("utf-8", errors="ignore") + _TRUNCATION_MARKER
        truncated = True

    return {
        "ok": True,
        "filename": filename,
        "content_type": content_type,
        "text": text,
        "truncated": truncated,
    }
